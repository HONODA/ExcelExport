import json
from shutil import copyfile
from shutil import SameFileError
import openpyxl
import re

class tool():

    db_address = 'Settings\\db.json'
    sql_address = 'Settings\\sql.json'
    template_address = 'templates\\对账单模板.xlsx'

    @staticmethod
    def loadJson():
        fi = open(tool.db_address, encoding='utf-8')  
        setting = json.load(fi)
        content = setting
        return content
    @staticmethod
    def loadSqlJson():
        f = open(tool.sql_address, encoding='utf-8')  
        setting = json.load(f)
        content = setting
        return content
    @staticmethod
    def Copy(path = ''):
        try:
            copyfile(tool.template_address,path)
        except IOError as e:
            return '拒绝访问'    #拒绝访问
        except SameFileError as e:
            return '存在相同文件'   #存在相同文件
        except Exception as e:
            return e.with_traceback()
    @staticmethod
    def replace_Excel_Argv(_argv,_list,_path,_title = '',_times = 1):
        """
        根据 _argv 的参数查找excel表中存在的 _argv 并替换成_content
        :param _list : 一个list
        :param _argv: 替换参数 eg: 在excel中存在 &argv1 
        :param _path: 是excel文件路径
        :param _title: WrokSheet 中的标题名称
        return 
        """
        try:
            wk = openpyxl.load_workbook(_path,False)
        except IOError as e:
            return "拒绝访问，检查是否以打开该文件："+e.args
        ws = wk.active
        if _title != '':
            ws.title = _title
        for i in ws.iter_cols():
            for value in i:
                if value.value != None and str(value.value).find(_argv) >= 0:
                    colnum = value.column
                    row = value.row
                    for j in _list :
                        rule = Rule()
                        mdict = rule.get_rulelist(value.value)
                        
                        for g in mdict.keys():
                            if g == '名称':
                               content = mdict.get(g,None)
                               continue
                            if g == '次数':
                                times = mdict.get(g,None)

                            tool.clear_argv(ws,row,colnum,Rule.rulelist[g],1) #清空除【名称】外的规则
                        for _i in range(len(content)):    #返回tuple
                            newcontent = j[content[_i]]
                            if times != None and len(times) != 0:
                                tool.replace_argv(ws,row,colnum,_argv,newcontent,times[_i])
                            else:
                                tool.replace_argv(ws,row,colnum,_argv,newcontent,_times)
                            _i +=1
                        
        wk.save(_path)
        wk.close()
        return "1"
    @staticmethod
    def replace_argv(_work_sheet,_row,_col,_old,_content,_times = 1):
            
            for i in range(int(_times)):
                mcell = _work_sheet.cell(_row + i,_col)
                if mcell.value != "" and mcell.value != None:
                    mcell.value = mcell.value.replace(_old,str(_content))
                else:
                    mcell.value = _content
    @staticmethod
    def clear_argv(_work_sheet,_row,_col,_argv,_times = 1):
            
            for i in range(int(_times)):
                mcell = _work_sheet.cell(_row + i,_col)
                if mcell.value != "" and mcell.value != None:
                    mcell.value = re.sub(_argv + '(.*)'+ _argv,'',mcell.value,int(_times))
                
        

    @staticmethod
    def rule_check():
        pass
class Rule(object):

    rulelist = {'名称':'&','次数':'@times@'}

    @staticmethod
    def add_rule(_key,_value):
        Rule.rulelist[_key,_value]
    @staticmethod
    def return_search(_argv,_str,_islist=False):
        '''
            根据所给参数返回查询到的中间内容
            :param islist 为 False 则返回为tuple
        '''
        mlist = re.findall(_argv + '(.*)'+ _argv,_str)
        if _islist :
            return mlist
        else:
            return tuple(mlist)
 
    def get_rulelist(self,_str):
        rlist ={}
        for i in Rule.rulelist:
            v = Rule.return_search(Rule.rulelist[i],_str)
            rlist[i] = v
        return rlist
    